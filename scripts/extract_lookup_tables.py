"""
Extract lookup tables and data from Excel workbooks.

This script reads the HC2-Calculators.xlsm file and extracts:
- Named ranges
- Lookup tables from the "Lookups" sheet
- Equipment specifications
- Material properties
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def extract_named_ranges(workbook_path: Path, output_dir: Path):
    """Extract all named ranges from workbook"""
    print("Loading workbook...")
    wb = load_workbook(workbook_path, data_only=True)

    named_ranges = {}

    print(f"Extracting named ranges...")
    for name in wb.defined_names.definedName:
        try:
            # Get the name and its reference
            range_name = name.name
            destinations = list(name.destinations)

            if destinations:
                sheet_name, cell_ref = destinations[0]
                ws = wb[sheet_name]

                # Handle single cell vs range
                if ':' in cell_ref:
                    # Range of cells
                    values = []
                    for row in ws[cell_ref]:
                        row_values = [cell.value for cell in row]
                        values.append(row_values)
                    named_ranges[range_name] = {
                        'sheet': sheet_name,
                        'reference': cell_ref,
                        'value': values,
                        'type': 'range'
                    }
                else:
                    # Single cell
                    cell = ws[cell_ref]
                    named_ranges[range_name] = {
                        'sheet': sheet_name,
                        'reference': cell_ref,
                        'value': cell.value,
                        'type': 'cell'
                    }
        except Exception as e:
            print(f"  Warning: Could not extract '{name.name}': {e}")

    # Save to JSON
    output_file = output_dir / 'named_ranges.json'
    with open(output_file, 'w') as f:
        json.dump(named_ranges, f, indent=2, default=str)

    print(f"Extracted {len(named_ranges)} named ranges to {output_file}")
    return named_ranges


def extract_lookups_sheet(workbook_path: Path, output_dir: Path):
    """Extract data from the Lookups sheet"""
    print("Extracting Lookups sheet...")

    try:
        # Read the Lookups sheet
        df = pd.read_excel(workbook_path, sheet_name='Lookups', header=None)

        # Save as CSV for easy editing
        output_file = output_dir / 'lookups_sheet.csv'
        df.to_csv(output_file, index=False, header=False)

        print(f"Saved Lookups sheet to {output_file}")

        # Try to parse structured tables from the sheet
        # (This would need customization based on actual layout)
        return df

    except Exception as e:
        print(f"Could not extract Lookups sheet: {e}")
        return None


def extract_fluid_properties(workbook_path: Path, output_dir: Path):
    """Extract fluid property data"""
    print("Extracting fluid properties...")

    wb = load_workbook(workbook_path, data_only=True)

    # Look for fluid-related named ranges
    fluid_data = {}

    # Common fluid names from the VBA analysis
    fluids = ['Water', 'Thermal Oil', 'Dowtherm A', 'Dowtherm J', 'Glycol']

    # This is a placeholder - actual implementation would need to
    # identify where fluid data is stored in the Excel file
    for fluid in fluids:
        fluid_data[fluid] = {
            'name': fluid,
            'properties': 'To be extracted from Excel'
        }

    output_file = output_dir / 'fluid_properties.json'
    with open(output_file, 'w') as f:
        json.dump(fluid_data, f, indent=2)

    print(f"Saved fluid properties template to {output_file}")


def extract_equipment_data(workbook_path: Path, output_dir: Path):
    """Extract equipment specifications"""
    print("Extracting equipment data...")

    # Read Item Lookup and Item Table sheets if they exist
    sheets_to_extract = ['Item Lookup', 'Item Table']

    for sheet_name in sheets_to_extract:
        try:
            df = pd.read_excel(workbook_path, sheet_name=sheet_name)
            output_file = output_dir / f"{sheet_name.replace(' ', '_').lower()}.csv"
            df.to_csv(output_file, index=False)
            print(f"  Saved {sheet_name} to {output_file}")
        except Exception as e:
            print(f"  Could not extract {sheet_name}: {e}")


def extract_all_data(workbook_path: Path, output_dir: Path):
    """Extract all data from Excel workbook"""
    print(f"\nExtracting data from {workbook_path.name}\n")
    print("=" * 60)

    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    # Extract named ranges
    named_ranges = extract_named_ranges(workbook_path, output_dir)

    # Extract lookup sheet
    extract_lookups_sheet(workbook_path, output_dir)

    # Extract fluid properties
    extract_fluid_properties(workbook_path, output_dir)

    # Extract equipment data
    extract_equipment_data(workbook_path, output_dir)

    # Generate summary report
    summary = {
        'workbook': workbook_path.name,
        'named_ranges_count': len(named_ranges),
        'extraction_date': pd.Timestamp.now().isoformat(),
        'named_ranges_by_type': {
            'cell': sum(1 for nr in named_ranges.values() if nr['type'] == 'cell'),
            'range': sum(1 for nr in named_ranges.values() if nr['type'] == 'range'),
        }
    }

    summary_file = output_dir / 'extraction_summary.json'
    with open(summary_file, 'w') as f:
        json.dump(summary, f, indent=2)

    print("\n" + "=" * 60)
    print(f"Extraction complete!")
    print(f"Output directory: {output_dir}")
    print(f"Named ranges: {summary['named_ranges_count']}")


if __name__ == '__main__':
    # Paths
    repo_root = Path(__file__).parent.parent
    workbook_path = repo_root / 'sources' / 'HC2-Calculators.xlsm'
    output_dir = repo_root / 'data' / 'lookup_tables'

    # Extract all data
    extract_all_data(workbook_path, output_dir)
